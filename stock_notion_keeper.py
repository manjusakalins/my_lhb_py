# -*- coding: utf-8 -*-
import urllib2
import urllib
import simplejson as json
import time
import os
import socket
import random
import threading
import common_api

import os
import pickle
import codecs
import sys
import xlrd
from openpyxl import load_workbook
import xlsxwriter
import re
import operator
import random
sheet_col_len=0;
g_notion={};

def openpyxl_get_row_array(sheet, row_index):
	res=[]
	cur_row=(tuple(sheet.get_squared_range(1,row_index+1,50, row_index+1)))[0];
	#print len(cur_row);
	#print cur_row[0].value
	for i in range(0,openpyxl_get_column_num(sheet)):
		res.append(cur_row[i].value)
	return res

def openpyxl_get_column_num(sheet):
	global sheet_col_len;
	cur_row=(tuple(sheet.get_squared_range(1,1,50, 1)))[0];
	if sheet_col_len == 0:
		for i in range(sheet.get_highest_column()):
			if cur_row[i].value:
				continue;
			else:
				break;
		sheet_col_len=i;
	return sheet_col_len;

def stock_get_notion_ths(code):
	#curl 'http://stockpage.10jqka.com.cn/300104/'
	renotion=""
	cmd_url = 'curl --connect-timeout 5 -m 5 http://stockpage.10jqka.com.cn/' + str(code) + '/'
	web=os.popen(cmd_url).readlines();
	for idx in range(len(web)):
		line=web[idx];
		if line.find("<dt>涉及概念：</dt>") != -1:
			lines=web[idx+1];#+web[idx+2];
			lines=lines.split("\"");
			#print lines[1];
			renotion=lines[1];
			break;
	return renotion;

#rewrite notion xlsx when not got it in xlsx.
def __stock_get_notion(code):
	global g_notion;
	sheet_name="notion";
	kfile = "stock_notion.xlsx"
	rb = load_workbook(kfile);
	rs = rb[sheet_name];

	g_notion={};
	#not find get it from web and rewrite xlsx
	notion=stock_get_notion_ths(code)
	if notion == "":
		notion=stock_get_notion_ths(code)
	#rewrite xlsx
	tmpfile = "bk_stock_notion.xlsx"
	wb = xlsxwriter.Workbook(tmpfile)
	ws = wb.add_worksheet(sheet_name);
	for row_idx in range(0, rs.max_row):
		cur_row = openpyxl_get_row_array(rs, row_idx);
		if cur_row == []:
			break;
		ws.write(row_idx, 0, cur_row[0]);
		ws.write(row_idx, 1, cur_row[1]);
		g_notion[cur_row[0]] = cur_row[1];
	#print row_idx;
#	if row_idx != 0:
	row_idx = row_idx + 1;
	ws.write(row_idx, 0, code);
	ws.write(row_idx, 1, notion.decode("utf-8"));
	g_notion[code] = notion.decode("utf-8");
	wb.close();
	cmd = '''mv %s %s''' % (tmpfile,kfile)
	list_file = os.popen(cmd).readlines();
	return notion;
	
#load xlsx pre
def stock_get_notion_load_all():
	sheet_name="notion";
	kfile = "stock_notion.xlsx"
	rb = load_workbook(kfile);
	rs = rb[sheet_name];
	global g_notion;

	g_notion={};
	for row_idx in range(0, rs.max_row):
		cur_row = openpyxl_get_row_array(rs, row_idx);
		g_notion[cur_row[0]] = cur_row[1];

#out api
def stock_get_notion(code):
	global g_notion;
	if len(g_notion) == 0:
		stock_get_notion_load_all();
	for key in g_notion.keys():
		if code.find(key) != -1:
			return g_notion[key];
	return __stock_get_notion(code);
